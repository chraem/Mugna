from os import path, getcwd
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, Fill, PatternFill

def generate_stub(company_name, employees, deductions, period, prepared_by, ws):
    col_per_stub= 10
    row_per_stub= 22

    # Font
    ANB18= Font(name="Arial Narrow", size=18, bold=True)
    ANB12= Font(name="Arial Narrow", size=12, bold=True)
    ANN12= Font(name="Arial Narrow", size=12, bold=False)

    # Fill Color
    fill1= PatternFill(fill_type="solid", start_color="FFF2CC")
    fill2= PatternFill(fill_type="solid", start_color="FFE699")

    # Border Style
    TD= Side(border_style = "dashDotDot")
    TN= Side(border_style= "thin")
    MN= Side(border_style= "medium")

    # Border Position
    TDMN_TB= Border(top=TD, bottom=MN)
    TDMN_BT= Border(bottom=TD, top=MN)
    TDMN_LR= Border(left=TD, right=MN)
    TDMN_RL= Border(right=TD, left=MN)
    TD_TL= Border(top=TD, left=TD)
    TD_BL= Border(bottom=TD, left=TD)
    TD_TR= Border(top=TD, right=TD)
    TD_BR= Border(bottom=TD, right=TD)
    MN_TB= Border(top=MN, bottom=MN)
    MN_TBL= Border(top=MN, bottom=MN, left=MN)
    MN_TBR= Border(top=MN, bottom=MN, right=MN)
    TN_B= Border(bottom=TN)
    TN_TRB= Border(top=TN, right=TN, bottom=TN)
    TN_RB= Border(right=TN, bottom=TN)
    TN_R= Border(right=TN)

    # Alignment
    LC_HV= Alignment(horizontal="left", vertical="center")
    RC_HV= Alignment(horizontal="right", vertical="center")
    CC_HV= Alignment(horizontal="center", vertical="center")
    left= True
    right= skip= False

    # Shorter rows
    short_rows= [1, 3, 5, 7, 22]

    for emp, employee_data in enumerate(employees):
        # Outsets
        x_onset= emp % 2 * col_per_stub + 1
        y_onset= emp // 2 * row_per_stub + 1

        # Simulation of X Onset Computation # Simulation of Y Onset Computation
        # emp # %2 # *11 # + 1              # emp # //2 # *11 # + 1
        #  0  # 0  #  0  #  1               #  0  #  0  #  0  #  1
        #  1  # 1  #  11 #  12              #  1  #  0  #  0  #  1
        #  2  # 0  #  0  #  1               #  2  #  1  #  11 #  12
        #  3  # 1  #  11 #  12              #  3  #  1  #  11 #  12
        #  4  # 0  #  0  #  1               #  4  #  2  #  22 #  23

        x= [get_column_letter(x_onset + i) for i in range(col_per_stub)]
        x.insert(0,'-')
        # -, A, B, C, D, E, F, G, H, I, J
        # -, K, L, M, N, O, P, Q, R, S, T, U

        y= [str(row_number) for row_number in range(y_onset, row_per_stub + y_onset)]
        y.insert(0,'-')
        # emp # y_onset # row_per_stub + y_onset # row_number ends at
        #  0  # 1       # 22 + 1                 # 23
        #  1  # 1       # 22 + 1                 # 23
        #  2  # 12      # 22 + 12                # 34
        #  3  # 12      # 22 + 12                # 34
        #  4  # 23      # 22 + 23                # 45

        # Stub Labels
        ws[x[2]+y[2]]= company_name
        ws[x[2]+y[2]].font= ANB18

        ws[x[9]+y[4]]= "Period: "+period
        ws[x[9]+y[4]].font= ANN12;

        ws[x[2]+y[4]]= "Employee ID"
        ws[x[3]+y[4]]= employee_data[0]
        ws[x[2]+y[4]].font= ws[x[3]+y[4]].font= ANB12
        
        ws[x[2]+y[6]]= "First Name"
        ws[x[3]+y[6]]= employee_data[1]
        ws[x[2]+y[6]].font= ws[x[3]+y[6]].font= ANN12

        ws[x[5]+y[6]]= "Middle Name"
        ws[x[6]+y[6]]= employee_data[2]
        ws[x[5]+y[6]].font= ws[x[6]+y[6]].font=ANN12

        ws[x[8]+y[6]]= "Last Name"
        ws[x[9]+y[6]]= employee_data[3]
        ws[x[8]+y[6]].font= ws[x[9]+y[6]].font=ANN12

        ws[x[2]+y[8]]= "Day Shift Hourly Rate"
        ws[x[3]+y[8]]= employee_data[4]
        ws[x[2]+y[8]].font=  ws[x[3]+y[8]].font= ANN12

        ws[x[5]+y[8]]= "Hour(s) Worked"
        ws[x[6]+y[8]]= employee_data[8]
        ws[x[5]+y[8]].font= ws[x[6]+y[8]].font= ANN12

        ws[x[8]+y[8]]= "Total Day Shift Pay"
        ws[x[9]+y[8]]= employee_data[12]
        ws[x[8]+y[8]].font= ws[x[9]+y[8]].font= ANN12

        ws[x[2]+y[9]]= "Night Shift Hourly Rate"
        ws[x[3]+y[9]]= employee_data[5]
        ws[x[2]+y[9]].font= ws[x[3]+y[9]].font= ANN12

        ws[x[5]+y[9]]= "Hour(s) Worked"
        ws[x[6]+y[9]]= employee_data[9]
        ws[x[5]+y[9]].font= ws[x[6]+y[9]].font= ANN12

        ws[x[8]+y[9]]= "Total Night Shift Pay"
        ws[x[9]+y[9]]= employee_data[13]
        ws[x[8]+y[9]].font= ws[x[9]+y[9]].font= ANN12

        ws[x[2]+y[10]]= "Holiday Hourly Rate"
        ws[x[3]+y[10]]= employee_data[6]
        ws[x[2]+y[10]].font= ws[x[3]+y[10]].font= ANN12

        ws[x[5]+y[10]]= "Hour(s) Worked"
        ws[x[6]+y[10]]= employee_data[10]
        ws[x[5]+y[10]].font= ws[x[6]+y[10]].font= ANN12

        ws[x[8]+y[10]]= "Total Holiday Pay"
        ws[x[9]+y[10]]= employee_data[14]
        ws[x[8]+y[10]].font= ws[x[9]+y[10]].font= ANN12

        ws[x[2]+y[11]]= "Overtime Hourly Rate"
        ws[x[3]+y[11]]= employee_data[7]
        ws[x[2]+y[11]].font= ws[x[3]+y[11]].font= ANN12

        ws[x[5]+y[11]]= "Hour(s) Worked"
        ws[x[6]+y[11]]= employee_data[11]
        ws[x[5]+y[11]].font= ws[x[6]+y[11]].font= ANN12

        ws[x[8]+y[11]]= "Total Overtime Pay"
        ws[x[9]+y[11]]= employee_data[15]
        ws[x[8]+y[11]].font= ws[x[9]+y[11]].font= ANN12

        ws[x[5]+y[13]]= "Gross Pay"
        ws[x[6]+y[13]]= employee_data[16]
        ws[x[5]+y[13]].font= ws[x[6]+y[13]].font= ANB12

        ws[x[5]+y[14]]= "Total Deduction"
        ws[x[6]+y[14]]= employee_data[17]
        ws[x[5]+y[14]].font= ws[x[6]+y[14]].font= ANB12

        ws[x[8]+y[13]]= "NET PAY"
        ws[x[9]+y[13]]= employee_data[18]
        ws[x[8]+y[13]].font= ws[x[9]+y[13]].font= ANB12
        
        ws[x[2]+y[13]]= "Deduction"
        ws[x[2]+y[13]].font= ANB12

        ws[x[3]+y[13]]= "Amount"
        ws[x[2]+y[14]].font= ANB12

        fullname = "{} {} {}".format(
            (employee_data[3].strip()+"," if employee_data[3] != "" else ""),
            (employee_data[1].strip()),
            ("".join(i[0]+"." for i in employee_data[2].split())),
            )

        ws[x[5]+y[18]]= "Received by "+fullname
        ws[x[5]+y[18]].font= ANN12        

        ws[x[8]+y[18]]= "Prepared by "+prepared_by
        ws[x[8]+y[18]].font= ANN12

        deduction_row = 14
        status = 0
        for deduction in deductions:
            if deduction[0] == employee_data[0]:
                ws[x[2]+y[deduction_row]]= deduction[1]
                ws[x[3]+y[deduction_row]]= deduction[2]
                deduction_row += 1
                status = 1
            elif deduction[0] != employee_data[0] and status == 1:
                break

        # Fills
        for col in range(2, col_per_stub):
            ws[x[col]+y[8]].fill= fill1
            ws[x[col]+y[10]].fill= fill1

        for col in range(2, 4):
            ws[x[col]+y[13]].fill= fill1

        for col in range(5, col_per_stub-3):
            ws[x[col]+y[13]].fill= fill1
            ws[x[col]+y[14]].fill= fill1

        for col in range(col_per_stub-2, col_per_stub): 
            ws[x[col]+y[13]].fill= fill2
            ws[x[col]+y[14]].fill= fill2

        # Outer Table Border
        if emp == 0 or emp == 1:
            for col in range(1, col_per_stub):
                ws[x[col]+'1'].border= TDMN_TB

        for col in range(1, col_per_stub):
            ws[x[col]+y[row_per_stub]].border= TDMN_BT
            ws[x[col]+y[1]].border= TDMN_TB

        for row in range(1, row_per_stub):
            ws[x[1]+y[row]].border= TDMN_LR
            ws[x[col_per_stub]+y[row]].border= TDMN_RL

        # Outer Table Corners
        ws[x[1]+y[1]].border= TD_TL
        ws[x[1]+y[row_per_stub]].border= TD_BL
        ws[x[col_per_stub]+y[1]].border= TD_TR
        ws[x[col_per_stub]+y[row_per_stub]].border= TD_BR

        # Inner Medium Border
        for col in range(3, col_per_stub-1):
            ws[x[col]+y[4]].border= MN_TB
        
        # Inner Medium Corners
        ws[x[2]+y[4]].border= MN_TBL
        ws[x[col_per_stub-1]+y[4]].border= MN_TBR

        # Inner Thin Border
        for col in range(2, col_per_stub):
            ws[x[col]+y[7]].border= TN_B
            ws[x[col]+y[11]].border= TN_B

        # Deduction Table
        for col in range(2, 4):
            ws[x[col]+y[13]].border= TN_TRB
            ws[x[col]+y[20]].border= TN_RB
            for row in range(14, 20):
                ws[x[col]+y[row]].border= TN_R

        # Signature Line
        ws.merge_cells(x[5]+y[19]+":"+x[6]+y[19])
        ws.merge_cells(x[8]+y[19]+":"+x[9]+y[19])
        ws[x[5]+y[19]].border= ws[x[6]+y[19]].border= TN_B
        ws[x[8]+y[19]].border= ws[x[9]+y[19]].border= TN_B

        # Net Pay
        ws.merge_cells(x[8]+y[13]+":"+x[8]+y[14])
        ws.merge_cells(x[9]+y[13]+":"+x[9]+y[14])

        # Alignment
        for col in range(2, col_per_stub):
            if left == True and skip == False:
                for row in range(1, row_per_stub):
                    ws[x[col]+y[row]].alignment= LC_HV
                right= True; left= False; skip= False
            elif right == True:
                for row in range(1, row_per_stub):
                    ws[x[col]+y[row]].alignment= RC_HV
                left= True; right= False; skip= True
            elif skip == True:
                continue

        ws[x[2]+y[13]].alignment= ws[x[3]+y[13]].alignment= CC_HV

        # Row width
        for row in short_rows:
            ws.row_dimensions[int(y[row])].height= 5

def save_wb(workbook, filename):
	workbook.save(path.normpath(getcwd()+"/mugna/exports/"+filename))
